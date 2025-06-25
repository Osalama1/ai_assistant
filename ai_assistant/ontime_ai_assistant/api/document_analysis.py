import frappe
import os
import PyPDF2
import docx
import openpyxl
from PIL import Image
import pytesseract
import io
from ai_assistant.ontime_ai_assistant.api.ai_service import get_ai_response
import pdfplumber

# Placeholder for ongoing processing tasks (consider using Frappe.cache or a DocType for persistence)
processing_tasks = {}

@frappe.whitelist()
def upload_document(file_url, document_name, document_type):
    try:
        # Ensure the file_url is a valid Frappe file URL
        if not file_url or not file_url.startswith("/files/"):
            frappe.throw("Invalid file URL provided. Must be a Frappe file URL (e.g., /files/my_document.pdf).")

        # Frappe stores private files in \'private/files/\' and public in \'public/files/\'
        # We need to determine the actual path on the server.
        # Assuming file_url comes from Frappe\'s standard upload, it will be in /files/ or /private/files/
        # frappe.get_site_path automatically handles public/private paths if the URL is correct.
        file_path = frappe.get_site_path(file_url)
        
        if not os.path.exists(file_path):
            frappe.throw(f"File not found at path: {file_path}")

        extracted_content = ""
        if document_type == "PDF":
            extracted_content = extract_text_from_pdf(file_path)
        elif document_type == "Word Document":
            extracted_content = extract_text_from_docx(file_path)
        elif document_type == "Excel Spreadsheet":
            extracted_content = extract_text_from_xlsx(file_path)
        elif document_type == "Image":
            extracted_content = extract_text_from_image(file_path)
        else:
            frappe.throw(f"Unsupported document type: {document_type}")

        # Get AI Settings and Provider details
        ai_settings = frappe.get_single("AI Settings")
        default_ai_provider_name = ai_settings.default_ai_provider
        ai_provider = frappe.get_doc("AI Provider", default_ai_provider_name)
        api_key = ai_provider.api_key

        # Send content to AI for analysis (as a background task)
        # Using frappe.enqueue for background processing
        task_id = frappe.enqueue(
            "ai_assistant.ontime_ai_assistant.api.document_analysis.analyze_document_with_ai",
            timeout=600, # 10 minutes timeout
            job_id=frappe.generate_hash(file_url + frappe.utils.now_datetime().isoformat()), # Unique job ID
            is_async=True,
            file_content=extracted_content,
            document_name=document_name,
            ai_provider_name=default_ai_provider_name,
            api_key=api_key
        )
        # Store task status in a temporary dictionary. For production, consider a DocType or Redis cache.
        processing_tasks[task_id] = {"status": "Processing", "extracted_data": None, "error": None}

        return {"success": True, "processor_id": task_id, "message": "Document submitted for analysis. You can check status using get_document_analysis_status.", "file_url": file_url}

    except Exception as e:
        frappe.log_error(f"Error in upload_document: {e}", "Document Analysis Error")
        return {"success": False, "error": str(e)}

def analyze_document_with_ai(file_content, document_name, ai_provider_name, api_key):
    # This function runs in a background job
    job_id = frappe.get_rq_job_id() # Get the job ID for updating status
    
    try:
        prompt = f"Analyze the following document content from \'{document_name}\' and extract key information, summarize it, and identify any relevant entities. If it\'s a structured document like an invoice or purchase order, extract line items, totals, dates, and parties. If it\'s a contract, identify key clauses, parties, and terms. If it\'s a general text, provide a concise summary and main topics. Return the output in a structured JSON format if possible, otherwise as a comprehensive summary.\n\nDocument Content:\n{file_content}"
        
        ai_response = get_ai_response(prompt, "Document Analysis", ai_provider_name, api_key)
        
        # Attempt to parse as JSON, otherwise keep as string
        try:
            extracted_data = frappe.parse_json(ai_response)
        except:
            extracted_data = ai_response

        processing_tasks[job_id] = {"status": "Completed", "extracted_data": extracted_data, "error": None}
    except Exception as e:
        frappe.log_error(f"Error analyzing document with AI: {e}", "Document Analysis Error")
        processing_tasks[job_id] = {"status": "Failed", "extracted_data": None, "error": str(e)}

@frappe.whitelist()
def get_processing_status(processor_id):
    status = processing_tasks.get(processor_id, {"status": "Unknown", "extracted_data": None, "error": "Processor ID not found or job not started."})
    return {"success": True, "status": status["status"], "extracted_data": status["extracted_data"], "error": status["error"]}

def extract_text_from_pdf(file_path):
    text = ""
    try:
        with pdfplumber.open(file_path) as pdf:
            for page in pdf.pages:
                text += page.extract_text() or ""
    except Exception as e:
        frappe.log_error(f"Error extracting text from PDF {file_path}: {e}", "Document Extraction Error")
        text = f"Error extracting text from PDF: {e}"
    return text

def extract_text_from_docx(file_path):
    text = ""
    try:
        doc = docx.Document(file_path)
        for para in doc.paragraphs:
            text += para.text + "\n"
    except Exception as e:
        frappe.log_error(f"Error extracting text from DOCX {file_path}: {e}", "Document Extraction Error")
        text = f"Error extracting text from DOCX: {e}"
    return text

def extract_text_from_xlsx(file_path):
    text = ""
    try:
        workbook = openpyxl.load_workbook(file_path)
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            text += f"\n--- Sheet: {sheet_name} ---\n"
            for row in sheet.iter_rows():
                row_values = [str(cell.value) if cell.value is not None else "" for cell in row]
                text += "\t".join(row_values) + "\n"
    except Exception as e:
        frappe.log_error(f"Error extracting text from XLSX {file_path}: {e}", "Document Extraction Error")
        text = f"Error extracting text from XLSX: {e}"
    return text

def extract_text_from_image(file_path):
    text = ""
    try:
        # Ensure Tesseract is installed and configured in the environment
        # For Frappe, you might need to ensure tesseract-ocr is installed on the server
        image = Image.open(file_path)
        text = pytesseract.image_to_string(image, lang=\'eng+ara\') # Support English and Arabic
    except Exception as e:
        frappe.log_error(f"Error extracting text from image {file_path}: {e}", "Document Extraction Error")
        text = f"Error extracting text from image: {e}"
    return text


